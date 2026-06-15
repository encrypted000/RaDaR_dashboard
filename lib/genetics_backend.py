"""
Genetics Backend Pipeline

Purpose
-------
This pipeline identifies RADAR patients who are expected to have a genetics
report but currently do not.

Steps
-----
1. Connect to RADAR database through SSH tunnel
2. Run SQL query to retrieve patients missing biopsy reports
3. Compare current results with the previous snapshot
4. Identify:
      - NEW_MISSING
      - STILL_MISSING
      - RESOLVED
5. Save outputs:
      - Raw snapshot CSV (used by Streamlit)
      - Excel report (for manual review)
      - Weekly KPI history

Run
---
python -m scripts.run_biopsy_backend
"""

# =============================================================================
# IMPORTS
# =============================================================================

from datetime import date, timedelta, datetime
from pathlib import Path

import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

from lib.db import get_connection
from lib.queries import SQL_GENETICS_MISSING, SQL_GENETICS_TOTAL_ELIGIBLE


# =============================================================================
# SETTINGS
# =============================================================================

MODULE = "genetics"


# =============================================================================
# PROJECT PATHS
# =============================================================================

def project_root() -> Path:
    """Return project root directory."""
    return Path(__file__).resolve().parents[1]


def module_dirs(module: str = MODULE) -> dict[str, Path]:
    """
    Create and return output directories for this module.

    outputs/<module>/
        raw/        → snapshot CSVs used by Streamlit
        reports/    → Excel reports for humans
        kpis/       → weekly KPI history
    """

    base = project_root() / "outputs" / module

    raw = base / "raw"
    reports = base / "reports"
    kpis = base / "kpis"

    raw.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)
    kpis.mkdir(parents=True, exist_ok=True)

    return {
        "base": base,
        "raw": raw,
        "reports": reports,
        "kpis": kpis
    }


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def week_monday_key(d: date | None = None) -> str:
    """Return Monday date (YYYY-MM-DD) for the current week."""
    d = d or date.today()
    monday = d - timedelta(days=d.weekday())
    return monday.isoformat()


def get_previous_snapshot_csv(raw_dir: Path, module: str = MODULE) -> Path | None:
    """Return the previous snapshot CSV (second most recent), not the current one.

    On each run, the pipeline:
      1. Reads prev_ids from this file   ← must be the OLD snapshot
      2. Saves a NEW snapshot file
      3. Computes diff between new and prev_ids

    If we returned files[0] (most recent), a second run on the same day would
    compare the brand-new snapshot against itself → new_missing=0, resolved=0.

    By returning files[1] (second most recent), we always compare against the
    snapshot from the previous run, giving correct diffs.

    Falls back to files[0] if only one snapshot exists (very first run).
    """
    files = sorted(raw_dir.glob(f"missing_{module}_*.csv"), reverse=True)
    if not files:
        return None
    # If 2+ snapshots exist, use second most recent as "previous"
    # so the most recent doesn't get compared against itself
    return files[1] if len(files) >= 2 else files[0]


# =============================================================================
# DATABASE QUERIES
# =============================================================================

def run_missing_query(conn) -> pd.DataFrame:
    """
    Run biopsy missing query and return dataframe.
    """

    df = pd.read_sql(SQL_GENETICS_MISSING, conn)

    df["patient_id"] = df["patient_id"].astype(str)
    df["date_of_birth"] = pd.to_datetime(df["date_of_birth"], errors="coerce")
    df["diagnosed_date"] = pd.to_datetime(df["diagnosed_date"], errors="coerce")
    df["recruited_date"] = pd.to_datetime(df["recruited_date"], errors="coerce")

    return df


def run_scalar_query(conn, sql: str) -> int:
    """
    Run a scalar SQL query returning a single numeric value.
    Used for KPIs such as total eligible patients.
    """

    cur = conn.cursor()
    cur.execute(sql)

    result = cur.fetchone()
    val = result[0] if result is not None else 0

    cur.close()

    return int(val or 0)


# =============================================================================
# STATUS COMPUTATION
# =============================================================================

def compute_status(df_missing: pd.DataFrame, prev_ids: set[str]):
    """
    Compare current snapshot with previous snapshot.

    Returns
    -------
    dataframe with status column
    current ids
    new ids
    still missing ids
    resolved ids
    """

    this_ids = set(df_missing["patient_id"].astype(str))

    new_ids = this_ids - prev_ids
    still_ids = this_ids & prev_ids
    resolved_ids = prev_ids - this_ids

    out = df_missing.copy()

    out["status"] = out["patient_id"].apply(
        lambda x: "NEW_MISSING" if x in new_ids else "STILL_MISSING"
    )

    return out, this_ids, new_ids, still_ids, resolved_ids


# =============================================================================
# EXCEL REPORT
# =============================================================================

def _col_width(ws, col_idx: int, min_w: int = 10, max_w: int = 40):
    """Auto-fit column width based on content."""
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[letter]:
        try:
            max_len = max(max_len, len(str(cell.value or "")))
        except Exception:
            pass
    ws.column_dimensions[letter].width = min(max(max_len + 3, min_w), max_w)


def write_excel_report(
    df_missing: pd.DataFrame,
    resolved_ids: set[str],
    runstamp: str,
    reports_dir: Path,
    module: str = MODULE,
    total_eligible: int = 0,
    uploaded: int = 0,
    completeness: float = 0.0,
    new_ids: set | None = None,
    still_ids: set | None = None,
    adult_missing: int = 0,
    child_missing: int = 0,
) -> Path:
    """
    Create a clean, professional Excel report with:
      Sheet 1 — Summary        (KPI dashboard with colour-coded metrics)
      Sheet 2 — Missing_Now    (all missing patients, colour-coded by status)
      Sheet 3 — New_Missing    (only newly appeared patients this week)
      Sheet 4 — Resolved       (patients no longer missing)
    """
    from openpyxl.styles import Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    new_ids   = new_ids   or set()
    still_ids = still_ids or set()

    excel_path = reports_dir / f"{module}_missing_report_{runstamp}.xlsx"
    wb = Workbook()

    # ── Shared styles ────────────────────────────────────────────────────────
    NAVY    = "0F2557"
    WHITE   = "FFFFFF"
    RED_BG  = "FFC7CE"
    RED_FG  = "9C0006"
    AMBER_BG= "FFEB9C"
    AMBER_FG= "9C6500"
    GREEN_BG= "C6EFCE"
    GREEN_FG= "276221"
    BLUE_BG = "DDEEFF"
    GREY_BG = "F2F2F2"
    STRIPE  = "F8FAFC"

    thin = Side(style="thin", color="D0D9EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(size=11, color=WHITE, bold=True):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def body_font(size=10, bold=False, color="1E293B"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def centre():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def left():
        return Alignment(horizontal="left", vertical="center")

    # ── SHEET 1: SUMMARY ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = f"RaDaR · {module.capitalize()} Missing Report"
    ws_sum["A1"].font    = Font(name="Calibri", size=16, bold=True, color=WHITE)
    ws_sum["A1"].fill    = fill(NAVY)
    ws_sum["A1"].alignment = centre()
    ws_sum.row_dimensions[1].height = 36

    ws_sum.merge_cells("A2:D2")
    ws_sum["A2"] = f"Generated: {runstamp.replace('_', ' ')}"
    ws_sum["A2"].font      = Font(name="Calibri", size=10, color="94A3B8")
    ws_sum["A2"].fill      = fill("1E3A8A")
    ws_sum["A2"].alignment = centre()
    ws_sum.row_dimensions[2].height = 20

    ws_sum.row_dimensions[3].height = 10

    kpi_headers = ["Metric", "Value", "Details", "Status"]
    for ci, h in enumerate(kpi_headers, start=1):
        cell = ws_sum.cell(row=4, column=ci, value=h)
        cell.font      = hdr_font(size=10, color=WHITE)
        cell.fill      = fill("1A3A8F")
        cell.alignment = centre()
        cell.border    = border
    ws_sum.row_dimensions[4].height = 22

    missing_total = df_missing["patient_id"].nunique() if not df_missing.empty else 0
    comp_status = "✅ Good" if completeness >= 75 else ("⚠ Needs attention" if completeness >= 50 else "❌ Low")

    kpi_rows = [
        ("Total Eligible Patients",  total_eligible,              "Registered in RaDaR",          BLUE_BG),
        ("Reports Uploaded",         uploaded,                    "Successfully submitted",        GREEN_BG),
        ("Missing Reports",          missing_total,               "Outstanding this week",         RED_BG  if missing_total > 0 else GREEN_BG),
        ("New Missing (this week)",  len(new_ids),                "Appeared since last run",       RED_BG  if len(new_ids) > 0  else GREEN_BG),
        ("Still Missing",            len(still_ids),              "Carried over from last run",    AMBER_BG),
        ("Resolved (this week)",     len(resolved_ids),           "No longer missing",             GREEN_BG),
        ("Adult Missing",            adult_missing,               "Age ≥ 18",                      BLUE_BG),
        ("Child Missing",            child_missing,               "Age < 18",                      BLUE_BG),
        ("Completeness",             f"{completeness:.1f}%",      comp_status,                     GREEN_BG if completeness >= 75 else (AMBER_BG if completeness >= 50 else RED_BG)),
    ]

    for ri, (metric, value, detail, bg) in enumerate(kpi_rows, start=5):
        ws_sum.cell(row=ri, column=1, value=metric).font = body_font(bold=True)
        ws_sum.cell(row=ri, column=1).fill = fill(GREY_BG)
        ws_sum.cell(row=ri, column=1).alignment = left()
        ws_sum.cell(row=ri, column=1).border = border

        ws_sum.cell(row=ri, column=2, value=value).font = body_font(bold=True, size=11)
        ws_sum.cell(row=ri, column=2).fill = fill(bg)
        ws_sum.cell(row=ri, column=2).alignment = centre()
        ws_sum.cell(row=ri, column=2).border = border

        ws_sum.cell(row=ri, column=3, value=detail).font = body_font(color="64748B")
        ws_sum.cell(row=ri, column=3).fill = fill(STRIPE)
        ws_sum.cell(row=ri, column=3).alignment = left()
        ws_sum.cell(row=ri, column=3).border = border

        if bg == RED_BG:
            status_txt, status_bg = "⚠ Action needed", RED_BG
        elif bg == GREEN_BG:
            status_txt, status_bg = "✓ OK", GREEN_BG
        else:
            status_txt, status_bg = "→ Monitor", AMBER_BG
        ws_sum.cell(row=ri, column=4, value=status_txt).font = body_font(bold=True)
        ws_sum.cell(row=ri, column=4).fill = fill(status_bg)
        ws_sum.cell(row=ri, column=4).alignment = centre()
        ws_sum.cell(row=ri, column=4).border = border
        ws_sum.row_dimensions[ri].height = 20

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 28
    ws_sum.column_dimensions["D"].width = 18

    # ── SHEET 2: MISSING NOW ─────────────────────────────────────────────────
    def write_patient_sheet(wb, title, df, status_filter=None,
                             hdr_color="1A3A8F", new_color=RED_BG, still_color=AMBER_BG):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False

        if df.empty:
            ws["A1"] = "No records."
            return ws

        if status_filter:
            df = df[df["status"] == status_filter].copy() if "status" in df.columns else df.copy()

        if df.empty:
            ws["A1"] = "No records."
            return ws

        col_labels = {
            "patient_id":       "Patient ID",
            "date_of_birth":    "Date of Birth",
            "age":              "Age",
            "diagnosed_date":   "Diagnosed",
            "recruited_date":   "Recruited",
            "hospital_name":    "Hospital",
            "status":           "Status",
        }
        display_cols = [c for c in col_labels if c in df.columns]
        headers = [col_labels[c] for c in display_cols]

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = hdr_font(size=10)
            cell.fill      = fill(hdr_color)
            cell.alignment = centre()
            cell.border    = border
        ws.row_dimensions[1].height = 22

        for ri, (_, row) in enumerate(df[display_cols].iterrows(), start=2):
            status_val = row.get("status", "") if "status" in row else ""
            row_bg = new_color if status_val == "NEW_MISSING" else still_color

            for ci, col in enumerate(display_cols, start=1):
                val = row[col]
                if col in ("date_of_birth", "diagnosed_date", "recruited_date"):
                    try:
                        val = pd.to_datetime(val).strftime("%d %b %Y") if pd.notna(val) else ""
                    except Exception:
                        val = str(val) if pd.notna(val) else ""
                elif col == "age":
                    val = int(val) if pd.notna(val) else ""

                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = body_font()
                cell.fill      = fill(row_bg)
                cell.alignment = left()
                cell.border    = border

            # Zebra stripe override for still missing
            if status_val != "NEW_MISSING" and ri % 2 == 0:
                for ci in range(1, len(display_cols) + 1):
                    ws.cell(row=ri, column=ci).fill = fill("FFF8E8")

            ws.row_dimensions[ri].height = 18

        ws.freeze_panes = "A2"

        for ci in range(1, len(display_cols) + 1):
            _col_width(ws, ci)

        note_row = ws.max_row + 2
        ws.cell(row=note_row, column=1, value=f"Total: {len(df)} patient(s)").font = body_font(bold=True, color="64748B")

        return ws

    write_patient_sheet(wb, "Missing_Now",   df_missing, hdr_color="7B1A1A", new_color=RED_BG,   still_color=AMBER_BG)
    write_patient_sheet(wb, "New_Missing",   df_missing, status_filter="NEW_MISSING",   hdr_color="7B1A1A", new_color=RED_BG,   still_color=RED_BG)
    write_patient_sheet(wb, "Still_Missing", df_missing, status_filter="STILL_MISSING", hdr_color="9C6500", new_color=AMBER_BG, still_color=AMBER_BG)

    # ── SHEET 5: RESOLVED ────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resolved")
    ws_res.sheet_view.showGridLines = False

    ws_res.merge_cells("A1:B1")
    ws_res["A1"] = "Patients Resolved This Week"
    ws_res["A1"].font      = hdr_font(size=11)
    ws_res["A1"].fill      = fill("276221")
    ws_res["A1"].alignment = centre()
    ws_res.row_dimensions[1].height = 24

    ws_res.cell(row=2, column=1, value="Patient ID").font = hdr_font(size=10, color=WHITE)
    ws_res.cell(row=2, column=1).fill = fill("1A3A8F")
    ws_res.cell(row=2, column=1).alignment = centre()
    ws_res.cell(row=2, column=1).border = border
    ws_res.cell(row=2, column=2, value="Note").font = hdr_font(size=10, color=WHITE)
    ws_res.cell(row=2, column=2).fill = fill("1A3A8F")
    ws_res.cell(row=2, column=2).alignment = centre()
    ws_res.cell(row=2, column=2).border = border

    for ri, pid in enumerate(sorted(resolved_ids), start=3):
        ws_res.cell(row=ri, column=1, value=pid).font = body_font()
        ws_res.cell(row=ri, column=1).fill = fill(GREEN_BG if ri % 2 == 0 else "EAFBEA")
        ws_res.cell(row=ri, column=1).alignment = left()
        ws_res.cell(row=ri, column=1).border = border
        ws_res.cell(row=ri, column=2, value="Report now on file").font = body_font(color="276221")
        ws_res.cell(row=ri, column=2).fill = fill(GREEN_BG if ri % 2 == 0 else "EAFBEA")
        ws_res.cell(row=ri, column=2).alignment = left()
        ws_res.cell(row=ri, column=2).border = border
        ws_res.row_dimensions[ri].height = 18

    if not resolved_ids:
        ws_res.cell(row=3, column=1, value="No patients resolved this week.").font = body_font(color="94A3B8")

    ws_res.freeze_panes = "A3"
    ws_res.column_dimensions["A"].width = 20
    ws_res.column_dimensions["B"].width = 24

    wb.save(excel_path)
    return excel_path


def update_excel_summary(*args, **kwargs):
    """No-op — summary is now written directly in write_excel_report."""
    pass


# =============================================================================
# KPI HISTORY
# =============================================================================

def update_weekly_kpis(
    kpis_dir: Path,
    week_monday: str,
    total_eligible: int,
    uploaded: int,
    missing_total: int,
    completeness_percent: float,
    new_missing: int,
    still_missing: int,
    resolved: int,
    adult_missing: int,
    child_missing: int,
):

    kpi_path = kpis_dir / "weekly_kpis.csv"

    row = pd.DataFrame([{
        "week_monday": week_monday,
        "missing_count": missing_total,
        "new_missing": new_missing,
        "still_missing": still_missing,
        "resolved": resolved,
        "total_eligible": total_eligible,
        "uploaded": uploaded,
        "completeness_percent": completeness_percent,
        "adult_missing": adult_missing,
        "child_missing": child_missing,
    }])

    if kpi_path.exists():

        existing = pd.read_csv(kpi_path)

        same_week = existing[existing["week_monday"] == week_monday]

        if not same_week.empty:
            # Row for this week already exists — preserve new_missing and resolved
            # from the FIRST run (when prev snapshot was genuinely different).
            # On re-runs within the same week, new_missing/resolved would be 0
            # because we now correctly point at the previous file.
            # Keep whichever run had the higher new_missing (i.e. the real diff).
            existing_new_missing = same_week.iloc[-1].get("new_missing", 0)
            existing_resolved    = same_week.iloc[-1].get("resolved", 0)
            if int(new_missing) == 0 and int(existing_new_missing) > 0:
                row["new_missing"] = existing_new_missing
            if int(resolved) == 0 and int(existing_resolved) > 0:
                row["resolved"] = existing_resolved

        existing = existing[existing["week_monday"] != week_monday]

        combined = pd.concat([existing, row], ignore_index=True)

        combined.to_csv(kpi_path, index=False)

    else:

        row.to_csv(kpi_path, index=False)


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_backend():
    """
    Run biopsy pipeline.
    """

    dirs = module_dirs(MODULE)

    raw_dir = dirs["raw"]
    reports_dir = dirs["reports"]
    kpis_dir = dirs["kpis"]

    prev_csv = get_previous_snapshot_csv(raw_dir)

    prev_ids = set()

    if prev_csv and prev_csv.exists():

        prev = pd.read_csv(prev_csv, dtype={"patient_id": "string"})

        prev_ids = set(prev["patient_id"].astype(str))

    conn, tunnel = get_connection()

    try:

        df_missing = run_missing_query(conn)

        total_eligible = run_scalar_query(conn, SQL_GENETICS_TOTAL_ELIGIBLE)

    finally:

        conn.close()
        tunnel.stop()

    missing_total = df_missing["patient_id"].nunique()

    uploaded = max(0, total_eligible - missing_total)  # guard against negative

    completeness = round(uploaded / total_eligible * 100, 2) if total_eligible else 0

    runstamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    df_missing2, this_ids, new_ids, still_ids, resolved_ids = compute_status(df_missing, prev_ids)

    if "age" in df_missing2.columns:

        adult_missing = df_missing2[df_missing2["age"] >= 18]["patient_id"].nunique()

        child_missing = df_missing2[df_missing2["age"] < 18]["patient_id"].nunique()

        # Patients with unknown age — count them as adults to avoid undercounting
        unknown_age = df_missing2[df_missing2["age"].isna()]["patient_id"].nunique()
        if unknown_age > 0:
            print(f"  ⚠  {unknown_age} patients have no age recorded — counted as adult")
            adult_missing += unknown_age

    else:

        adult_missing = 0
        child_missing = 0

    snap_csv = raw_dir / f"missing_{MODULE}_{runstamp}.csv"

    _expected_cols = ["patient_id", "date_of_birth", "age", "diagnosed_date", "recruited_date", "hospital_name", "hospital_code", "status"]
    _save_cols = [col for col in _expected_cols if col in df_missing2.columns]
    df_missing2[_save_cols].to_csv(snap_csv, index=False)

    excel_path = write_excel_report(
        df_missing2, resolved_ids, runstamp, reports_dir,
        total_eligible=total_eligible,
        uploaded=uploaded,
        completeness=completeness,
        new_ids=new_ids,
        still_ids=still_ids,
        adult_missing=adult_missing,
        child_missing=child_missing,
    )


    update_weekly_kpis(
        kpis_dir,
        week_monday_key(),
        total_eligible,
        uploaded,
        missing_total,
        completeness,
        len(new_ids),
        len(still_ids),
        len(resolved_ids),
        adult_missing,
        child_missing,
    )

    # Keep only the 12 most recent snapshots and reports (avoid folder bloat)
    _snaps = sorted(raw_dir.glob(f"missing_{MODULE}_*.csv"), reverse=True)
    for _old in _snaps[12:]:
        _old.unlink()

    _reports = sorted(reports_dir.glob(f"{MODULE}_missing_report_*.xlsx"), reverse=True)
    for _old in _reports[12:]:
        _old.unlink()

    print("=" * 50)
    print(f"  Module          : {MODULE}")
    print(f"  Total eligible  : {total_eligible}")
    print(f"  Uploaded        : {uploaded}")
    print(f"  Missing now     : {missing_total}")
    print(f"  Completeness    : {completeness:.1f}%")
    print(f"  New missing     : {len(new_ids)}")
    print(f"  Still missing   : {len(still_ids)}")
    print(f"  Resolved        : {len(resolved_ids)}")
    print(f"  Adult missing   : {adult_missing}")
    print(f"  Child missing   : {child_missing}")
    print("=" * 50)
    print("Saved raw snapshot:", snap_csv)
    print("Saved Excel report:", excel_path)
    print("Updated KPI:", kpis_dir / "weekly_kpis.csv")